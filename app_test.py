import unittest
from openpyxl.reader.excel import load_workbook
from requests.api import get
from weather import *
import openpyxl


class TestApp(unittest.TestCase):

    def test_ciudad_correcta(self):
        ciudad = 'saltillo'
        resultado = get_weather(ciudad)
        self.assertEqual(resultado, 'Exito')

    def test_ciudad_error(self):
        ciudad = 'saltilo'
        resultado = get_weather(ciudad)
        self.assertEqual(resultado, "Error encontrando clima para la ciudad '" + ciudad + "'")

    def test_bd_ingles(self):
        excel_doc = openpyxl.load_workbook('Libro1.xlsx')
        hoja = excel_doc.get_sheet_by_name('Hoja1')
        ciudad = hoja['A3'].value
        resultado = get_weather(ciudad)
        self.assertEqual(resultado, 'Exito')
        
    def test_bd_esp(self):
        excel_doc = openpyxl.load_workbook('Libro1.xlsx')
        hoja = excel_doc.get_sheet_by_name('Hoja1')
        ciudad = hoja['A4'].value
        resultado = get_weather(ciudad)
        self.assertEqual(resultado, 'Exito')
    
    def test_bd_error(self):
        excel_doc = openpyxl.load_workbook('Libro1.xlsx')
        hoja = excel_doc.get_sheet_by_name('Hoja1')
        ciudad = hoja['A5'].value
        resultado = get_weather(ciudad)
        self.assertEqual(resultado, "Error encontrando clima para la ciudad '" + ciudad + "'")

    def test_bd_min_may(self):
        excel_doc = openpyxl.load_workbook('Libro1.xlsx')
        hoja = excel_doc.get_sheet_by_name('Hoja1')
        ciudad = hoja['A6'].value
        resultado = get_weather(ciudad)
        self.assertEqual(resultado, "Error encontrando clima para la ciudad '" + ciudad + "'")


if __name__ == '__main__':
    unittest.main()